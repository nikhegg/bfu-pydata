import openpyxl

workbook = openpyxl.load_workbook("task.xlsx")
sheet = workbook.active

max_salary = {
    "Фамилия": "",
    "Сумма зарплаты, руб.": 0
}
min_salary = {
    "Фамилия": "",
    "Сумма зарплаты, руб.": float('inf')
}
dep_salaries = {}

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):
    if row[0].value and row[0].value != "Итог" and row[0].value != "Общий итог":
        salary = row[5].value
        if salary > max_salary["Сумма зарплаты, руб."]:
            max_salary = {
                "Фамилия": row[1].value,
                "Сумма зарплаты, руб.": salary
            }
        if salary < min_salary["Сумма зарплаты, руб."]:
            min_salary = {
                "Фамилия": row[1].value,
                "Сумма зарплаты, руб.": salary
            }

        department = row[2].value
        if department not in dep_salaries:
            dep_salaries[department] = {
                "total_salary": 0,
                "count": 0
            }
        dep_salaries[department]["total_salary"] += salary
        dep_salaries[department]["count"] += 1


department_averages = {}
for department, values in dep_salaries.items():
    department_averages[department] = values["total_salary"] / values["count"]

row_num = 14
sheet.cell(row=row_num, column=2, value="Человек с макс. зарплатой:")
sheet.cell(row=row_num, column=3, value=f"{max_salary['Фамилия']}: {max_salary['Сумма зарплаты, руб.']} руб.")
row_num += 1

sheet.cell(row=row_num, column=2, value="Человек с мин. зарплатой:")
sheet.cell(row=row_num, column=3, value=f"{min_salary['Фамилия']}: {min_salary['Сумма зарплаты, руб.']} руб.")
row_num += 1

sheet.cell(row=row_num, column=2, value="Средняя зарплата по отделам:")
for department, avg_salary in department_averages.items():
    row_num += 1
    sheet.cell(row=row_num, column=2, value=department)
    sheet.cell(row=row_num, column=3, value=f"{avg_salary:.2f} руб.")

workbook.save("task.xlsx")